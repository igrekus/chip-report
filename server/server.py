import json
import random

import openpyxl

from flask import Flask, request, send_file
from flask_cors import CORS, cross_origin
from openpyxl.styles import Alignment

app = Flask(__name__)
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'


@app.route('/')
@cross_origin()
def hello():
    return 'server ok'


@app.route('/api', methods=('GET', 'POST'))
@cross_origin()
def api():
    data = json.loads(request.data)

    wb = openpyxl.Workbook()
    ws = wb.active

    align_center = Alignment(wrapText=True, shrinkToFit=True, horizontal='center', vertical='center')

    starting_cell = ws['A1']
    table_offset = 0
    for table in data['tables']:
        header, indexes, norms = build_columns(table['columns'])
        header_cell = starting_cell.offset(table_offset, 0)

        # append header
        merge_row = header_cell.row
        merge_col_start = header_cell.column
        merge_col_end = merge_col_start + len(header) - 1
        ws.merge_cells(start_row=merge_row, start_column=merge_col_start, end_row=merge_row, end_column=merge_col_end)
        header_cell.value = table['header']
        header_cell.alignment = align_center

        for h, n, i in zip(header, norms, range(len(header))):
            cell = header_cell.offset(1, i)
            cell.value = h
            cell.alignment = align_center

            cell = header_cell.offset(2, i)
            cell.value = n
            cell.alignment = align_center

            cell = header_cell.offset(3, i)
            cell.value = i + 1
            cell.alignment = align_center

        # append generated data
        body_cell = header_cell.offset(4, 0)
        row_number = table['rows'] if table['rows'] else 0
        for row in range(row_number):
            row_data = generate_row(row, table['columns'])
            for col, r_d in enumerate(row_data):
                cell = body_cell.offset(row, col)
                cell.value = r_d
                cell.alignment = align_center

        table_offset += 4 + row_number

    wb.save('server\\tmp\\out.xlsx')
    return send_file('tmp\\out.xlsx', as_attachment=True)


def build_columns(columns: list):
    header_row = list()
    indexes = list()
    norms = list()

    header_row.append('№ изделия')
    indexes.append('name')
    norms.append('Норма')

    for col in columns:
        cond = '\n' + col['condition'] if col['condition'] else ''
        header_row.append(f'{col["label"]}{cond}')
        indexes.append(col['index'])
        norms.append(col['norms'])
    return header_row, indexes, norms


def generate_row(row_index, columns):
    return [row_index + 1] + [generate_value(col) for col in columns]


def generate_value(column):
    if column:
        spread, step, mid = column['spread'], column['step'], column['mid']
        if spread and step and mid:
            start = mid - spread
            stop = mid + spread
            return round(random.randint(0, int((stop - start) / step)) * step + start, 2)
        else:
            return mid
    else:
        return '-'
